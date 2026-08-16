#!/usr/bin/env python3
"""Build the Gate12BL truthful release package and formal supplementary tables."""

from __future__ import annotations

import csv
import hashlib
import importlib.metadata
import json
import os
import platform
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


RUNTIME_PYTHON = "python3"
COMMAND_PREFIX = f"{RUNTIME_PYTHON} scripts/build_gate12bl_reproducibility_release.py"

NEW_DATA_AVAILABILITY = (
    "All public input datasets analysed in this study are available from GEO under accession numbers "
    "GSE143791, GSE202813, GSE266330 and GSE323357, and from the National Omics Data Encyclopedia "
    "under OEP005136 (https://www.biosino.org/node/project/detail/OEP005136). The OEP005136 paired "
    "raw-matrix subset used for reconstruction was obtained from NODE analysis OEZ00021715, data object "
    "OED01122886 (Paired_mBone.zip). Figure-linked source-data tables are included in the accompanying "
    "submission package. The exact release-build audit configuration is included for provenance only. "
    "Analysis and rendering scripts, scientific-analysis configuration files and intermediate computational "
    "objects are not included in the current package. A public repository DOI or stable URL for those "
    "reproducibility materials has not yet been assigned and must be inserted before submission."
)

S4_RESULTS_ANCHOR = (
    "Stable computation was therefore not interpreted as support for three conserved biological ecotypes."
)
S4_RESULTS_SENTENCE = (
    "The representation-by-evidence-layer boundary for both frozen models is summarized in Supplementary Table S4."
)

S4_TITLE = (
    "Supplementary Table S4. Evidence boundaries for the frozen 81-gene T/NK program and "
    "three-ecotype representation."
)
S4_DESCRIPTION = (
    "Discovery stability, external directional support and external structure/agreement are reported as "
    "distinct, non-interchangeable evidence layers."
)
S4_HEADERS = ["Representation", "Evidence layer", "Status"]
S4_FOOTNOTES = [
    "T/NK, T/natural-killer.",
    "‘Discovery’ denotes derivation-stage evidence and does not constitute external validation.",
    "‘External direction’ denotes preservation of the prespecified score direction in independent datasets. "
    "The common 81-gene program did not preserve its discovery direction in GSE266330 or OEP005136.",
    "‘External structure/agreement’ denotes independently derived consensus structure and agreement with "
    "non-forced transferred labels in OEP005136.",
    "At the frozen k=3 evaluation, mean silhouette was 0.292, but the proportion of ambiguous clustering was "
    "0.464 and the minimum cluster size was 2; agreement with transferred labels was absent (adjusted Rand "
    "index, −0.004; 10,000-permutation P=0.492).",
    "‘Not supported’ means that a prespecified criterion was not met. ‘Not tested’ means that the evidence "
    "layer was not evaluated and must not be interpreted as evidence of absence.",
    "‘Stable intersection’ and ‘internally specified’ describe discovery-stage construction only and do not "
    "establish external transportability.",
]

S6_TITLE = "Supplementary Table S6. Claim-level evidence boundaries for the frozen Axis1 framework."
S6_DESCRIPTION = (
    "The strongest interpretation permitted by the technical reconstruction, paired human endpoint and linked "
    "spatial analyses is reported for each claim class."
)
S6_HEADERS = ["Claim", "Evidence status", "Evidence basis"]
S6_FOOTNOTES = [
    "‘Frozen technical reconstruction’ refers to projection without refitting. The ‘47 projectable external "
    "patients’ basis refers specifically to 47 of 49 OEP005136 patients; 46 of 47 GSE266330 donors or patients "
    "were projectable in the separate unpaired directional test.",
    "‘Linked spatial organization’ is based on four linked mouse sections, of which three were decision-evaluable "
    "under the primary geometry. Unique animal identifiers were unavailable; no animal-level population inference "
    "was made.",
    "The matched bone-metastasis-versus-normal-bone rule was not met: one of two patient pairs was positive, and "
    "only 5.8% of 499 within-specimen subsampling iterations retained the frozen rule.",
    "‘Bone specificity’ was not tested because no matched non-bone metastatic comparator was available.",
    "‘Clinical prediction’ was not tested because no prespecified protected clinical endpoint or independent "
    "outcome-validation set was available.",
    "‘Causal mechanism’ was not tested because no perturbational validation or virtual knockout was performed.",
    "‘Supported’ denotes passage of the specified operational endpoint only. It does not upgrade technical "
    "reconstruction or linked spatial organization into universal biological validation, specificity, prediction "
    "or causality.",
    "BM, bone metastasis.",
]

AUTHOR_CONTROLLED_TEXT = [
    "**Authors:** [AUTHOR LIST TO BE COMPLETED]",
    "**Affiliations:** [AFFILIATIONS TO BE COMPLETED]",
    "**Corresponding author:** [FULL CONTACT DETAILS TO BE COMPLETED]",
    "[FUNDING SOURCES, GRANT NUMBERS AND FUNDER ROLES TO BE COMPLETED. If no specific funding supported the work, this must be confirmed by all authors.]",
    "[AUTHOR CONFIRMATION REQUIRED. Suggested wording if accurate: “The authors declare no competing interests.”]",
    "[AUTHOR NAMES AND CRediT ROLES TO BE COMPLETED WITHOUT INVENTING CONTRIBUTIONS: Conceptualization, Data curation, Formal analysis, Funding acquisition, Investigation, Methodology, Project administration, Resources, Software, Supervision, Validation, Visualization, Writing – original draft, Writing – review and editing.]",
]

RESOLVE_IDS = {"BI-10", "BI-11", "BI-12", "BI-13"}
EXPECTED_INHERITED_RESOLVED = {"BI-08", "BI-09", "BI-16"}
EXPECTED_UNRESOLVED = {
    "BI-01", "BI-02", "BI-03", "BI-04", "BI-05", "BI-06", "BI-07", "BI-14", "BI-15", "BI-17"
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def write_tsv(path: Path, rows: Iterable[dict[str, Any]], fields: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle, delimiter="\t", fieldnames=list(fields), extrasaction="ignore", lineterminator="\n"
        )
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def load_config(path: Path) -> dict[str, str]:
    rows = read_tsv(path)
    if not rows or set(rows[0]) != {"key", "value"}:
        raise RuntimeError("Gate12BL configuration must have key and value columns")
    config: dict[str, str] = {}
    for row in rows:
        key = row["key"]
        if key in config:
            raise RuntimeError(f"Duplicate configuration key: {key}")
        config[key] = row["value"]
    return config


def all_files(path: Path) -> list[Path]:
    return sorted(item for item in path.rglob("*") if item.is_file())


def add_check(
    checks: list[dict[str, Any]],
    check_id: str,
    domain: str,
    observed: Any,
    expected: Any,
    passed: bool,
    note: str = "",
    severity: str = "HARD",
) -> None:
    checks.append(
        {
            "check_id": check_id,
            "domain": domain,
            "observed": observed,
            "expected": expected,
            "status": "PASS" if passed else ("WARN" if severity == "WARN" else "FAIL"),
            "severity": severity,
            "note": note,
        }
    )


def section(text: str, heading: str, next_heading: str | None) -> str:
    start = text.index(heading)
    end = len(text) if next_heading is None else text.index(next_heading, start + len(heading))
    return text[start:end]


def expand_citation(token: str) -> list[int]:
    if not re.fullmatch(r"\d+(?:(?:,|–|-)\d+)*", token):
        return []
    output: list[int] = []
    for item in token.split(","):
        match = re.fullmatch(r"(\d+)[–-](\d+)", item)
        if match:
            start, end = map(int, match.groups())
            output.extend(range(start, end + 1))
        else:
            output.append(int(item))
    return output


def citation_numbers(text: str) -> list[int]:
    body = text.split("# References\n", 1)[0]
    numbers: list[int] = []
    for token in re.findall(r"<sup>([^<]+)</sup>", body):
        numbers.extend(expand_citation(token))
    return numbers


def reference_records(text: str) -> list[str]:
    block = text.split("# References\n", 1)[1].split("# Figure legends\n", 1)[0]
    return [
        match.group(0).strip()
        for match in re.finditer(r"(?ms)^\d+\.\s+.*?(?=^\d+\.\s+|\Z)", block.strip())
    ]


def manifest_rows(root: Path, excluded: set[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in all_files(root):
        relative = path.relative_to(root).as_posix()
        if relative in excluded:
            continue
        rows.append({"relative_path": relative, "bytes": path.stat().st_size, "sha256": sha256_file(path)})
    return rows


def build_legends_markdown() -> str:
    def block(title: str, description: str, notes: list[str]) -> str:
        lines = [f"## {title}", "", description, "", "**Footnotes**", ""]
        lines.extend(f"{index}. {note}" for index, note in enumerate(notes, start=1))
        return "\n".join(lines)

    return (
        "# Supplementary table legends\n\n"
        + block(S4_TITLE, S4_DESCRIPTION, S4_FOOTNOTES)
        + "\n\n"
        + block(S6_TITLE, S6_DESCRIPTION, S6_FOOTNOTES)
        + "\n\n## Numbering note\n\n"
        + "The manuscript currently cites Supplementary Tables S4 and S6. Numbering gaps are intentionally "
        + "preserved until the target journal and complete supplementary-file convention are selected; no table "
        + "was silently renumbered in Gate12BL.\n"
    )


def style_table_sheet(
    worksheet: Any,
    title: str,
    description: str,
    headers: list[str],
    data: list[list[str]],
    footnotes: list[str],
) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    worksheet.cell(1, 1, title)
    worksheet.cell(1, 1).font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    worksheet.cell(1, 1).fill = PatternFill("solid", fgColor="333333")
    worksheet.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[1].height = 34

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    worksheet.cell(2, 1, description)
    worksheet.cell(2, 1).font = Font(name="Arial", size=9, italic=True, color="333333")
    worksheet.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[2].height = 31

    thin_gray = Side(style="thin", color="A6A6A6")
    for column, value in enumerate(headers, start=1):
        cell = worksheet.cell(4, column, value)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(top=thin_gray, bottom=thin_gray)
    worksheet.row_dimensions[4].height = 25

    for row_index, row in enumerate(data, start=5):
        for column, value in enumerate(row, start=1):
            cell = worksheet.cell(row_index, column, value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))
        worksheet.row_dimensions[row_index].height = 31

    footnote_header_row = 12
    worksheet.merge_cells(start_row=footnote_header_row, start_column=1, end_row=footnote_header_row, end_column=3)
    worksheet.cell(footnote_header_row, 1, "Footnotes")
    worksheet.cell(footnote_header_row, 1).font = Font(name="Arial", size=9, bold=True)
    worksheet.cell(footnote_header_row, 1).fill = PatternFill("solid", fgColor="E7E6E6")
    for offset, note in enumerate(footnotes, start=1):
        row_index = footnote_header_row + offset
        worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
        worksheet.cell(row_index, 1, f"{offset}. {note}")
        worksheet.cell(row_index, 1).font = Font(name="Arial", size=8)
        worksheet.cell(row_index, 1).alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.row_dimensions[row_index].height = 28 if len(note) < 150 else 42

    widths = [34, 32, 34]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width
    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = f"A4:C{4 + len(data)}"
    worksheet.print_area = f"A1:C{footnote_header_row + len(footnotes)}"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 1
    worksheet.page_margins = PageMargins(left=0.3, right=0.3, top=0.45, bottom=0.45, header=0.2, footer=0.2)
    worksheet.oddFooter.center.text = "Page &P of &N"


def build_workbook(path: Path, s4_data: list[list[str]], s6_data: list[list[str]], source_hashes: dict[str, str]) -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "Read me"
    readme.sheet_view.showGridLines = False
    readme.merge_cells("A1:C1")
    readme["A1"] = "Supplementary Tables S4 and S6"
    readme["A1"].font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    readme["A1"].fill = PatternFill("solid", fgColor="333333")
    readme["A1"].alignment = Alignment(vertical="center")
    readme.row_dimensions[1].height = 34
    readme_rows = [
        ("Purpose", "Journal-neutral formalization of the two supplementary tables cited in the active manuscript."),
        ("Data integrity", "The data regions are copied cell-for-cell from the frozen Gate12BH machine-readable TSV files."),
        ("Table S4 source SHA-256", source_hashes["s4"]),
        ("Table S6 source SHA-256", source_hashes["s6"]),
        ("Scope boundary", "No formula, macro, hidden data, external link or color-dependent conclusion is used."),
        ("Numbering", "S4 and S6 are retained exactly as cited. Gaps are a target-journal formatting warning; no table was renumbered."),
        ("Evidence boundary", "Not tested is not evidence of absence; supported refers only to the specified operational endpoint."),
    ]
    for row_index, (key, value) in enumerate(readme_rows, start=3):
        readme.cell(row_index, 1, key).font = Font(name="Arial", size=9, bold=True)
        readme.cell(row_index, 2, value).font = Font(name="Arial", size=9)
        readme.cell(row_index, 1).alignment = Alignment(wrap_text=True, vertical="top")
        readme.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
        readme.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=3)
        readme.row_dimensions[row_index].height = 34 if len(value) > 100 else 24
    readme.column_dimensions["A"].width = 28
    readme.column_dimensions["B"].width = 48
    readme.column_dimensions["C"].width = 26
    readme.freeze_panes = "A3"
    readme.sheet_properties.pageSetUpPr.fitToPage = True
    readme.page_setup.orientation = "landscape"
    readme.page_setup.paperSize = readme.PAPERSIZE_A4
    readme.page_setup.fitToWidth = 1
    readme.page_setup.fitToHeight = 1
    readme.print_area = f"A1:C{2 + len(readme_rows)}"

    s4_sheet = workbook.create_sheet("Table S4")
    style_table_sheet(s4_sheet, S4_TITLE, S4_DESCRIPTION, S4_HEADERS, s4_data, S4_FOOTNOTES)
    s6_sheet = workbook.create_sheet("Table S6")
    style_table_sheet(s6_sheet, S6_TITLE, S6_DESCRIPTION, S6_HEADERS, s6_data, S6_FOOTNOTES)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_pdf(path: Path, s4_data: list[list[str]], s6_data: list[list[str]]) -> None:
    font_paths = {
        "GateArial": Path("/System/Library/Fonts/Supplemental/Arial.ttf"),
        "GateArialBold": Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
        "GateArialItalic": Path("/System/Library/Fonts/Supplemental/Arial Italic.ttf"),
    }
    if all(font_path.is_file() for font_path in font_paths.values()):
        for font_name, font_path in font_paths.items():
            if font_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
        pdfmetrics.registerFontFamily(
            "GateArial",
            normal="GateArial",
            bold="GateArialBold",
            italic="GateArialItalic",
            boldItalic="GateArialBold",
        )
        regular_font, bold_font, italic_font = "GateArial", "GateArialBold", "GateArialItalic"
    else:
        regular_font, bold_font, italic_font = "Helvetica", "Helvetica-Bold", "Helvetica-Oblique"
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TableTitle", parent=styles["Heading2"], fontName=bold_font, fontSize=11, leading=13, spaceAfter=5
    )
    description_style = ParagraphStyle(
        "Description", parent=styles["BodyText"], fontName=italic_font, fontSize=7.8, leading=9.5, spaceAfter=6
    )
    note_style = ParagraphStyle(
        "Note", parent=styles["BodyText"], fontName=regular_font, fontSize=6.8, leading=8.2, alignment=TA_LEFT, spaceAfter=1.5
    )
    cell_style = ParagraphStyle(
        "Cell", parent=styles["BodyText"], fontName=regular_font, fontSize=7.4, leading=9, alignment=TA_LEFT
    )
    header_style = ParagraphStyle(
        "Header", parent=cell_style, fontName=bold_font, fontSize=7.5, leading=9
    )
    document = SimpleDocTemplate(
        str(path), pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title="Supplementary Tables S4 and S6",
        author="Gate12BL reproducibility-aligned release",
    )
    story: list[Any] = []

    def add_table(title: str, description: str, headers: list[str], rows: list[list[str]], notes: list[str]) -> None:
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(description, description_style))
        content = [[Paragraph(value, header_style) for value in headers]]
        content.extend([[Paragraph(str(value), cell_style) for value in row] for row in rows])
        table = Table(content, colWidths=[82 * mm, 82 * mm, 90 * mm], repeatRows=1, hAlign="LEFT")
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9D9D9")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#A6A6A6")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 5))
        story.append(Paragraph("<b>Footnotes</b>", note_style))
        for index, note in enumerate(notes, start=1):
            story.append(Paragraph(f"{index}. {note}", note_style))

    add_table(S4_TITLE, S4_DESCRIPTION, S4_HEADERS, s4_data, S4_FOOTNOTES)
    story.append(PageBreak())
    add_table(S6_TITLE, S6_DESCRIPTION, S6_HEADERS, s6_data, S6_FOOTNOTES)
    document.build(story)


def validate_workbook(
    path: Path,
    expected_s4: list[list[str]],
    expected_s6: list[list[str]],
    legends_text: str,
) -> dict[str, Any]:
    with zipfile.ZipFile(path, "r") as archive:
        corrupt_member = archive.testzip()
        member_count = len(archive.namelist())
    workbook = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    sheet_names = workbook.sheetnames
    s4_sheet = workbook["Table S4"]
    s6_sheet = workbook["Table S6"]
    observed_s4 = [[str(s4_sheet.cell(row, column).value) for column in range(1, 4)] for row in range(5, 11)]
    observed_s6 = [[str(s6_sheet.cell(row, column).value) for column in range(1, 4)] for row in range(5, 11)]
    s4_workbook_text = [str(s4_sheet["A1"].value), str(s4_sheet["A2"].value)] + [
        str(s4_sheet.cell(12 + index, 1).value) for index in range(1, len(S4_FOOTNOTES) + 1)
    ]
    s6_workbook_text = [str(s6_sheet["A1"].value), str(s6_sheet["A2"].value)] + [
        str(s6_sheet.cell(12 + index, 1).value) for index in range(1, len(S6_FOOTNOTES) + 1)
    ]
    formulas = []
    hidden_rows = []
    hidden_columns = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    formulas.append(f"{worksheet.title}!{cell.coordinate}")
        hidden_rows.extend(f"{worksheet.title}!{index}" for index, value in worksheet.row_dimensions.items() if value.hidden)
        hidden_columns.extend(
            f"{worksheet.title}!{index}" for index, value in worksheet.column_dimensions.items() if value.hidden
        )
    external_links = len(getattr(workbook, "_external_links", []))
    worksheet_states = {worksheet.title: worksheet.sheet_state for worksheet in workbook.worksheets}
    markdown_title_note_presence = all(
        token in legends_text
        for token in [S4_TITLE, S6_TITLE, *S4_FOOTNOTES, *S6_FOOTNOTES]
    )
    workbook_title_note_presence = (
        s4_workbook_text[:2] == [S4_TITLE, S4_DESCRIPTION]
        and s6_workbook_text[:2] == [S6_TITLE, S6_DESCRIPTION]
        and s4_workbook_text[2:] == [f"{index}. {note}" for index, note in enumerate(S4_FOOTNOTES, start=1)]
        and s6_workbook_text[2:] == [f"{index}. {note}" for index, note in enumerate(S6_FOOTNOTES, start=1)]
    )
    return {
        "zip_corrupt_member": corrupt_member,
        "zip_member_count": member_count,
        "sheet_names": sheet_names,
        "worksheet_states": worksheet_states,
        "s4_match": observed_s4 == expected_s4,
        "s6_match": observed_s6 == expected_s6,
        "s4_structure": (
            s4_sheet["A11"].value is None
            and s4_sheet["A12"].value == "Footnotes"
            and s4_sheet.max_row == 12 + len(S4_FOOTNOTES)
        ),
        "s6_structure": (
            s6_sheet["A11"].value is None
            and s6_sheet["A12"].value == "Footnotes"
            and s6_sheet.max_row == 12 + len(S6_FOOTNOTES)
        ),
        "formulas": formulas,
        "hidden_rows": hidden_rows,
        "hidden_columns": hidden_columns,
        "external_links": external_links,
        "titles_and_notes_in_markdown": markdown_title_note_presence,
        "titles_and_notes_in_workbook": workbook_title_note_presence,
    }


def validate_libreoffice(workbook_path: Path) -> dict[str, Any]:
    soffice = shutil.which("soffice")
    if not soffice:
        return {
            "available": False,
            "return_code": None,
            "pdf_pages": 0,
            "pdf_text": "",
            "stderr": "soffice not found",
        }
    with tempfile.TemporaryDirectory(prefix="gate12bl_soffice_") as temp_dir:
        profile_uri = (Path(temp_dir) / "libreoffice_profile").resolve().as_uri()
        completed = subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile_uri}",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                temp_dir,
                str(workbook_path),
            ],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
        converted = Path(temp_dir) / (workbook_path.stem + ".pdf")
        converted_reader = PdfReader(str(converted)) if converted.is_file() else None
        pages = len(converted_reader.pages) if converted_reader else 0
        converted_text = (
            "\n".join(page.extract_text() or "" for page in converted_reader.pages)
            if converted_reader
            else ""
        )
        return {
            "available": True,
            "return_code": completed.returncode,
            "pdf_pages": pages,
            "pdf_text": converted_text,
            "stderr": completed.stderr.strip(),
        }


def run_gate() -> int:
    started = datetime.now().astimezone()
    if len(sys.argv) != 2:
        raise RuntimeError(f"Usage: {COMMAND_PREFIX} <config.tsv>")

    root = Path.cwd().resolve()
    config_path = Path(sys.argv[1]).resolve()
    config_display = config_path.relative_to(root).as_posix()
    run_command = f"{COMMAND_PREFIX} {config_display}"
    config = load_config(config_path)
    required = {
        "source_bk_manuscript", "source_bk_legends", "source_bk_receipt", "source_bk_blockers",
        "source_bh_package", "source_bh_receipt", "source_bh_payload_manifest", "source_bh_figure_manifest",
        "source_s4", "source_s6", "gate_plan", "audit_script", "output_dir",
        "expected_bk_manuscript_sha256", "expected_bk_legends_sha256", "expected_bk_receipt_sha256",
        "expected_bk_blockers_sha256", "expected_bh_receipt_sha256", "expected_bh_payload_manifest_sha256",
        "expected_bh_figure_manifest_sha256", "expected_s4_sha256", "expected_s6_sha256",
        "expected_gate_plan_sha256", "expected_audit_script_sha256", "expected_figures",
        "expected_source_data_files", "expected_s4_rows", "expected_s6_rows", "expected_references",
        "expected_resolved_blockers", "expected_remaining_blockers", "expected_blocking_remaining",
        "expected_target_dependent_remaining",
    }
    missing = sorted(required - set(config))
    if missing:
        raise RuntimeError("Missing Gate12BL configuration keys: " + ", ".join(missing))

    paths = {
        "bk_manuscript": root / config["source_bk_manuscript"],
        "bk_legends": root / config["source_bk_legends"],
        "bk_receipt": root / config["source_bk_receipt"],
        "bk_blockers": root / config["source_bk_blockers"],
        "bh_receipt": root / config["source_bh_receipt"],
        "bh_payload_manifest": root / config["source_bh_payload_manifest"],
        "bh_figure_manifest": root / config["source_bh_figure_manifest"],
        "s4": root / config["source_s4"],
        "s6": root / config["source_s6"],
        "gate_plan": root / config["gate_plan"],
        "audit_script": root / config["audit_script"],
    }
    expected_hashes = {
        "bk_manuscript": config["expected_bk_manuscript_sha256"],
        "bk_legends": config["expected_bk_legends_sha256"],
        "bk_receipt": config["expected_bk_receipt_sha256"],
        "bk_blockers": config["expected_bk_blockers_sha256"],
        "bh_receipt": config["expected_bh_receipt_sha256"],
        "bh_payload_manifest": config["expected_bh_payload_manifest_sha256"],
        "bh_figure_manifest": config["expected_bh_figure_manifest_sha256"],
        "s4": config["expected_s4_sha256"],
        "s6": config["expected_s6_sha256"],
        "gate_plan": config["expected_gate_plan_sha256"],
        "audit_script": config["expected_audit_script_sha256"],
    }
    for role, path in paths.items():
        if not path.is_file():
            raise RuntimeError(f"Missing frozen Gate12BL input: {path}")
        observed = sha256_file(path)
        if observed != expected_hashes[role]:
            raise RuntimeError(f"Frozen input hash mismatch for {path}: expected {expected_hashes[role]}; observed {observed}")

    source_bh = (root / config["source_bh_package"]).resolve()
    if not source_bh.is_dir():
        raise RuntimeError(f"Missing Gate12BH package: {source_bh}")
    output = root / config["output_dir"]
    staging = output.parent / f".{output.name}.staging"
    if output.exists():
        raise RuntimeError(f"Refusing to overwrite existing Gate12BL output: {output}")
    if staging.exists():
        raise RuntimeError(f"Refusing to overwrite existing Gate12BL staging directory: {staging}")

    bk_receipt = json.loads(paths["bk_receipt"].read_text(encoding="utf-8"))
    bh_receipt = json.loads(paths["bh_receipt"].read_text(encoding="utf-8"))
    if bk_receipt.get("status") != "COMPLETED":
        raise RuntimeError("Gate12BK parent receipt is not COMPLETED")
    if bk_receipt.get("verdicts", {}).get("evidence_patch_integrity") != "PASS":
        raise RuntimeError("Gate12BK evidence patch did not pass")
    if bk_receipt.get("verdicts", {}).get("citation_reference_integrity") != "PASS":
        raise RuntimeError("Gate12BK citation/reference integrity did not pass")
    if bk_receipt.get("verdicts", {}).get("scientific_core_invariance") != "PASS":
        raise RuntimeError("Gate12BK scientific core invariance did not pass")
    if bh_receipt.get("status") != "COMPLETED":
        raise RuntimeError("Gate12BH parent receipt is not COMPLETED")
    if bh_receipt.get("audit", {}).get("total") != 80 or bh_receipt.get("audit", {}).get("passed") != 80:
        raise RuntimeError("Gate12BH parent audit is not 80/80 PASS")

    bh_manifest_rows = read_tsv(paths["bh_payload_manifest"])
    if not bh_manifest_rows or list(bh_manifest_rows[0]) != ["file", "bytes", "sha256"]:
        raise RuntimeError("Unexpected Gate12BH payload-manifest columns")
    bh_manifest = {row["file"]: row for row in bh_manifest_rows}
    source_figure_files = all_files(source_bh / "figures")
    source_data_files = all_files(source_bh / "source_data")
    if len(source_figure_files) != int(config["expected_figures"]):
        raise RuntimeError(f"Unexpected Gate12BH figure count: {len(source_figure_files)}")
    if len(source_data_files) != int(config["expected_source_data_files"]):
        raise RuntimeError(f"Unexpected Gate12BH source-data count: {len(source_data_files)}")
    for source_path in source_figure_files + source_data_files:
        relative = source_path.relative_to(source_bh).as_posix()
        record = bh_manifest.get(relative)
        if not record or record["sha256"] != sha256_file(source_path):
            raise RuntimeError(f"Gate12BH payload manifest mismatch: {relative}")

    s4_rows = read_tsv(paths["s4"])
    s6_rows = read_tsv(paths["s6"])
    if len(s4_rows) != int(config["expected_s4_rows"]):
        raise RuntimeError(f"Unexpected Table S4 row count: {len(s4_rows)}")
    if len(s6_rows) != int(config["expected_s6_rows"]):
        raise RuntimeError(f"Unexpected Table S6 row count: {len(s6_rows)}")
    if list(s4_rows[0]) != ["representation", "evidence_layer", "status"]:
        raise RuntimeError("Unexpected Table S4 source columns")
    if list(s6_rows[0]) != ["claim", "status", "basis"]:
        raise RuntimeError("Unexpected Table S6 source columns")
    s4_data = [[row["representation"], row["evidence_layer"], row["status"]] for row in s4_rows]
    s6_data = [[row["claim"], row["status"], row["basis"]] for row in s6_rows]

    staging.mkdir(parents=True)
    shutil.copytree(source_bh / "figures", staging / "figures", copy_function=shutil.copy2)
    shutil.copytree(source_bh / "source_data", staging / "source_data", copy_function=shutil.copy2)
    (staging / "manuscript").mkdir()
    (staging / "tables").mkdir()
    (staging / "admin").mkdir()
    (staging / "provenance").mkdir()
    frozen_config_path = staging / "provenance/gate12bl_reproducibility_release_v1.tsv"
    shutil.copy2(config_path, frozen_config_path)

    inherited_rows: list[dict[str, Any]] = []
    for source_path in source_figure_files + source_data_files:
        relative = source_path.relative_to(source_bh).as_posix()
        destination = staging / relative
        source_hash = sha256_file(source_path)
        output_hash = sha256_file(destination)
        inherited_rows.append(
            {
                "relative_path": relative,
                "source_bytes": source_path.stat().st_size,
                "output_bytes": destination.stat().st_size,
                "source_sha256": source_hash,
                "output_sha256": output_hash,
                "status": "PASS" if source_hash == output_hash else "FAIL",
            }
        )
    write_tsv(
        staging / "provenance/GATE12BL_INHERITED_PAYLOAD_AUDIT.tsv",
        inherited_rows,
        ["relative_path", "source_bytes", "output_bytes", "source_sha256", "output_sha256", "status"],
    )
    if any(row["status"] != "PASS" for row in inherited_rows):
        raise RuntimeError("Inherited Gate12BH payload changed during copy")

    source_manuscript = paths["bk_manuscript"].read_text(encoding="utf-8")
    source_legends = paths["bk_legends"].read_text(encoding="utf-8")
    old_data_section = section(source_manuscript, "# Data availability\n", "# Ethics statement\n")
    if old_data_section.count("Gate12BE review-driven submission package") != 1:
        raise RuntimeError("Gate12BK stale package statement was not found exactly once")
    revised_data_section = "# Data availability\n\n" + NEW_DATA_AVAILABILITY + "\n\n"
    revised_manuscript = source_manuscript.replace(old_data_section, revised_data_section, 1)
    if source_manuscript.count(S4_RESULTS_ANCHOR) != 1:
        raise RuntimeError("Gate12BK S4 results anchor was not found exactly once")
    if S4_RESULTS_SENTENCE in source_manuscript:
        raise RuntimeError("Gate12BK already contains the Gate12BL S4 cross-reference sentence")
    revised_manuscript = revised_manuscript.replace(
        S4_RESULTS_ANCHOR, S4_RESULTS_ANCHOR + " " + S4_RESULTS_SENTENCE, 1
    )
    if not source_legends.startswith("# Gate12BH figure legends\n"):
        raise RuntimeError("Unexpected inherited legend heading")
    revised_legends = source_legends.replace("# Gate12BH figure legends", "# Figure legends", 1)

    manuscript_path = staging / "manuscript/Gate12BL_Reproducibility_Aligned_Manuscript.md"
    legends_path = staging / "manuscript/Gate12BL_Figure_Legends.md"
    manuscript_path.write_text(revised_manuscript, encoding="utf-8")
    legends_path.write_text(revised_legends, encoding="utf-8")

    table_legends = build_legends_markdown()
    table_legends_path = staging / "tables/Supplementary_Table_Legends.md"
    table_legends_path.write_text(table_legends, encoding="utf-8")
    workbook_path = staging / "tables/Supplementary_Tables_S4_and_S6.xlsx"
    build_workbook(
        workbook_path,
        s4_data,
        s6_data,
        {"s4": expected_hashes["s4"], "s6": expected_hashes["s6"]},
    )
    pdf_path = staging / "tables/Supplementary_Tables_S4_and_S6.pdf"
    build_pdf(pdf_path, s4_data, s6_data)

    workbook_audit = validate_workbook(
        workbook_path,
        s4_data,
        s6_data,
        table_legends_path.read_text(encoding="utf-8"),
    )
    libreoffice_audit = validate_libreoffice(workbook_path)
    pdf_reader = PdfReader(str(pdf_path))
    pdf_text = "\n".join(page.extract_text() or "" for page in pdf_reader.pages)

    checks: list[dict[str, Any]] = []
    for role, path in paths.items():
        add_check(checks, f"HASH_{role.upper()}", "frozen-input", sha256_file(path), expected_hashes[role], True)
    add_check(checks, "BK_RECEIPT", "parent", bk_receipt.get("status"), "COMPLETED", True)
    add_check(checks, "BH_RECEIPT", "parent", bh_receipt.get("status"), "COMPLETED", True)
    add_check(checks, "BH_AUDIT", "parent", f"{bh_receipt['audit']['passed']}/{bh_receipt['audit']['total']}", "80/80", True)
    add_check(checks, "FIGURE_COUNT", "payload", len(source_figure_files), int(config["expected_figures"]), len(source_figure_files) == int(config["expected_figures"]))
    add_check(checks, "SOURCE_DATA_COUNT", "payload", len(source_data_files), int(config["expected_source_data_files"]), len(source_data_files) == int(config["expected_source_data_files"]))
    add_check(checks, "INHERITED_HASHES", "payload", sum(row["status"] == "PASS" for row in inherited_rows), len(inherited_rows), all(row["status"] == "PASS" for row in inherited_rows))

    new_data_section = section(revised_manuscript, "# Data availability\n", "# Ethics statement\n")
    gate_tokens = re.findall(r"Gate12[A-Za-z0-9_-]*", new_data_section)
    add_check(checks, "DATA_NO_INTERNAL_GATE", "data-availability", len(gate_tokens), 0, not gate_tokens)
    add_check(checks, "DATA_NO_OLD_CLAIM", "data-availability", revised_manuscript.count("review-driven submission package contains"), 0, "review-driven submission package contains" not in revised_manuscript)
    add_check(checks, "DATA_EXPLICIT_EXCLUSIONS", "data-availability", new_data_section.count("not included"), 1, new_data_section.count("not included") == 1)
    release_config_phrase = "The exact release-build audit configuration is included for provenance only."
    analysis_exclusion_phrase = "Analysis and rendering scripts, scientific-analysis configuration files and intermediate computational objects are not included in the current package."
    add_check(checks, "DATA_RELEASE_CONFIG_SCOPE", "data-availability", new_data_section.count(release_config_phrase), 1, new_data_section.count(release_config_phrase) == 1)
    add_check(checks, "DATA_ANALYSIS_MATERIAL_EXCLUSION", "data-availability", new_data_section.count(analysis_exclusion_phrase), 1, new_data_section.count(analysis_exclusion_phrase) == 1)
    add_check(checks, "RELEASE_CONFIG_INCLUDED", "payload", sha256_file(frozen_config_path), sha256_file(config_path), frozen_config_path.is_file() and sha256_file(frozen_config_path) == sha256_file(config_path))
    identifiers = [
        "GSE143791", "GSE202813", "GSE266330", "GSE323357", "OEP005136", "OEZ00021715",
        "OED01122886", "Paired_mBone.zip", "https://www.biosino.org/node/project/detail/OEP005136",
    ]
    missing_ids = [identifier for identifier in identifiers if identifier not in new_data_section]
    add_check(checks, "DATA_IDENTIFIERS", "data-availability", ",".join(missing_ids) or "all present", "all present", not missing_ids)
    add_check(checks, "DATA_SOURCE_TABLE_CLAIM", "data-availability", "included" if "Figure-linked source-data tables are included" in new_data_section else "absent", "included", "Figure-linked source-data tables are included" in new_data_section)
    add_check(checks, "DATA_REPOSITORY_NOT_FABRICATED", "data-availability", "not assigned" if "has not yet been assigned" in new_data_section else "missing", "not assigned", "has not yet been assigned" in new_data_section)

    source_results = section(source_manuscript, "# Results\n", "# Discussion\n")
    revised_results = section(revised_manuscript, "# Results\n", "# Discussion\n")
    masked_results = revised_results.replace(" " + S4_RESULTS_SENTENCE, "", 1)
    add_check(checks, "RESULTS_ONLY_S4_REFERENCE", "scientific-core", hashlib.sha256(masked_results.encode()).hexdigest(), hashlib.sha256(source_results.encode()).hexdigest(), masked_results == source_results)
    source_data_masked = revised_manuscript.replace(revised_data_section, old_data_section, 1).replace(" " + S4_RESULTS_SENTENCE, "", 1)
    add_check(checks, "MANUSCRIPT_TWO_PATCHES_ONLY", "scientific-core", hashlib.sha256(source_data_masked.encode()).hexdigest(), hashlib.sha256(source_manuscript.encode()).hexdigest(), source_data_masked == source_manuscript)
    restored_legends = revised_legends.replace("# Figure legends", "# Gate12BH figure legends", 1)
    add_check(checks, "LEGENDS_HEADING_ONLY", "scientific-core", hashlib.sha256(restored_legends.encode()).hexdigest(), hashlib.sha256(source_legends.encode()).hexdigest(), restored_legends == source_legends)
    placeholders_preserved = all(source_manuscript.count(item) == 1 and revised_manuscript.count(item) == 1 for item in AUTHOR_CONTROLLED_TEXT)
    add_check(checks, "AUTHOR_PLACEHOLDERS", "scope", "preserved" if placeholders_preserved else "changed", "preserved", placeholders_preserved)
    source_refs = reference_records(source_manuscript)
    revised_refs = reference_records(revised_manuscript)
    add_check(checks, "REFERENCE_COUNT", "citation", len(revised_refs), int(config["expected_references"]), len(revised_refs) == int(config["expected_references"]))
    add_check(checks, "REFERENCE_INVARIANCE", "citation", hashlib.sha256("\n".join(revised_refs).encode()).hexdigest(), hashlib.sha256("\n".join(source_refs).encode()).hexdigest(), revised_refs == source_refs)
    add_check(checks, "CITATION_NUMBER_INVARIANCE", "citation", citation_numbers(revised_manuscript), citation_numbers(source_manuscript), citation_numbers(revised_manuscript) == citation_numbers(source_manuscript))
    add_check(checks, "S4_MAIN_REFERENCE", "supplementary-table", revised_manuscript.count("Supplementary Table S4"), source_manuscript.count("Supplementary Table S4") + 1, revised_manuscript.count("Supplementary Table S4") == source_manuscript.count("Supplementary Table S4") + 1)
    add_check(checks, "S6_REFERENCE_INVARIANCE", "supplementary-table", revised_manuscript.count("Supplementary Table S6"), source_manuscript.count("Supplementary Table S6"), revised_manuscript.count("Supplementary Table S6") == source_manuscript.count("Supplementary Table S6"))
    add_check(checks, "S4_LEGEND_REFERENCE", "supplementary-table", revised_legends.count("Supplementary Table S4"), 1, revised_legends.count("Supplementary Table S4") == 1)
    add_check(checks, "S6_LEGEND_REFERENCE", "supplementary-table", revised_legends.count("Supplementary Table S6"), 1, revised_legends.count("Supplementary Table S6") == 1)
    s4_combinations = {(row["representation"], row["evidence_layer"]) for row in s4_rows}
    s6_claims = {row["claim"] for row in s6_rows}
    allowed_s6_statuses = {"Supported", "Supported with boundary", "Not established", "Not tested"}
    add_check(checks, "S4_UNIQUE_COMBINATIONS", "supplementary-table", len(s4_combinations), 6, len(s4_combinations) == 6)
    add_check(checks, "S6_UNIQUE_CLAIMS", "supplementary-table", len(s6_claims), 6, len(s6_claims) == 6)
    add_check(checks, "S6_STATUS_VOCABULARY", "supplementary-table", sorted({row["status"] for row in s6_rows}), sorted(allowed_s6_statuses), {row["status"] for row in s6_rows} <= allowed_s6_statuses)

    add_check(checks, "XLSX_ZIP", "supplementary-table", workbook_audit["zip_corrupt_member"], None, workbook_audit["zip_corrupt_member"] is None)
    add_check(checks, "XLSX_SHEETS", "supplementary-table", workbook_audit["sheet_names"], ["Read me", "Table S4", "Table S6"], workbook_audit["sheet_names"] == ["Read me", "Table S4", "Table S6"])
    add_check(checks, "XLSX_VISIBLE", "supplementary-table", workbook_audit["worksheet_states"], {"Read me": "visible", "Table S4": "visible", "Table S6": "visible"}, all(value == "visible" for value in workbook_audit["worksheet_states"].values()))
    add_check(checks, "XLSX_S4_DATA", "supplementary-table", workbook_audit["s4_match"], True, workbook_audit["s4_match"])
    add_check(checks, "XLSX_S6_DATA", "supplementary-table", workbook_audit["s6_match"], True, workbook_audit["s6_match"])
    add_check(checks, "XLSX_S4_STRUCTURE", "supplementary-table", workbook_audit["s4_structure"], True, workbook_audit["s4_structure"])
    add_check(checks, "XLSX_S6_STRUCTURE", "supplementary-table", workbook_audit["s6_structure"], True, workbook_audit["s6_structure"])
    add_check(checks, "XLSX_FORMULAS", "supplementary-table", len(workbook_audit["formulas"]), 0, not workbook_audit["formulas"])
    add_check(checks, "XLSX_HIDDEN_DATA", "supplementary-table", len(workbook_audit["hidden_rows"]) + len(workbook_audit["hidden_columns"]), 0, not workbook_audit["hidden_rows"] and not workbook_audit["hidden_columns"])
    add_check(checks, "XLSX_EXTERNAL_LINKS", "supplementary-table", workbook_audit["external_links"], 0, workbook_audit["external_links"] == 0)
    add_check(checks, "TABLE_TITLES_NOTES_MARKDOWN", "supplementary-table", workbook_audit["titles_and_notes_in_markdown"], True, workbook_audit["titles_and_notes_in_markdown"])
    add_check(checks, "TABLE_TITLES_NOTES_WORKBOOK", "supplementary-table", workbook_audit["titles_and_notes_in_workbook"], True, workbook_audit["titles_and_notes_in_workbook"])
    libreoffice_text = re.sub(r"\s+", " ", libreoffice_audit["pdf_text"])
    libreoffice_tokens = [
        "Supplementary Table S4", "Supplementary Table S6", "Stable intersection",
        "External structure/agreement", "Frozen technical reconstruction", "Causal mechanism",
    ]
    libreoffice_content_ok = all(token in libreoffice_text for token in libreoffice_tokens)
    add_check(checks, "LIBREOFFICE_OPEN", "supplementary-table", libreoffice_audit["return_code"], 0, libreoffice_audit["available"] and libreoffice_audit["return_code"] == 0 and libreoffice_audit["pdf_pages"] == 3 and libreoffice_content_ok, note=f"pages={libreoffice_audit['pdf_pages']}; content_ok={libreoffice_content_ok}; stderr={libreoffice_audit['stderr']}")
    add_check(checks, "PDF_PAGES", "supplementary-table", len(pdf_reader.pages), 2, len(pdf_reader.pages) == 2)
    normalized_pdf_text = re.sub(r"\s+", " ", pdf_text)
    pdf_tokens = [
        "Supplementary Table S4", "Supplementary Table S6", *S4_HEADERS, *S6_HEADERS,
        *[value for row in s4_data for value in row], *[value for row in s6_data for value in row],
        "1. T/NK, T/natural-killer.", "−0.004", "8. BM, bone metastasis.",
    ]
    missing_pdf_tokens = [token for token in pdf_tokens if token not in normalized_pdf_text]
    add_check(checks, "PDF_CONTENT", "supplementary-table", ", ".join(missing_pdf_tokens) or "complete", "complete", not missing_pdf_tokens)
    add_check(checks, "NUMBERING_GAP", "submission-format", "S4 and S6 only", "target-journal decision pending", False, note="SUPPLEMENTARY_TABLE_NUMBERING_GAP", severity="WARN")

    blockers = read_tsv(paths["bk_blockers"])
    blocker_rows: list[dict[str, Any]] = []
    for row in blockers:
        blocker_id = row["blocker_id"]
        updated = {
            ("gate12bi_resolved" if key == "resolved" else key): value
            for key, value in row.items()
        }
        if blocker_id in RESOLVE_IDS:
            updated["gate12bl_status"] = "RESOLVED"
            updated["gate12bl_resolved"] = "True"
            evidence = {
                "BI-10": "Active Data availability contains no internal Gate label.",
                "BI-11": "Active Data availability explicitly states that analysis and rendering scripts are not included.",
                "BI-12": "Active Data availability distinguishes the included release-build audit configuration from absent scientific-analysis configurations and intermediate computational objects.",
                "BI-13": "Formal XLSX/PDF Supplementary Tables S4/S6 and standalone table legends are present and audited against frozen TSVs.",
            }[blocker_id]
            updated["gate12bl_evidence"] = evidence
        elif row.get("gate12bk_status") == "RESOLVED":
            updated["gate12bl_status"] = "RESOLVED_INHERITED"
            updated["gate12bl_resolved"] = "True"
            updated["gate12bl_evidence"] = row.get("gate12bk_evidence", "Inherited from Gate12BK.")
        else:
            updated["gate12bl_status"] = "UNRESOLVED"
            updated["gate12bl_resolved"] = "False"
            updated["gate12bl_evidence"] = "Outside Gate12BL truthfulness/table-formalization scope."
        blocker_rows.append(updated)
    blocker_ids = {row["blocker_id"] for row in blocker_rows}
    if blocker_ids != EXPECTED_INHERITED_RESOLVED | RESOLVE_IDS | EXPECTED_UNRESOLVED:
        raise RuntimeError(f"Unexpected blocker universe: {sorted(blocker_ids)}")
    resolved_ids = {row["blocker_id"] for row in blocker_rows if row["gate12bl_status"] in {"RESOLVED", "RESOLVED_INHERITED"}}
    unresolved_rows = [row for row in blocker_rows if row["gate12bl_status"] == "UNRESOLVED"]
    unresolved_ids = {row["blocker_id"] for row in unresolved_rows}
    blocking_remaining = sum(row["severity"] == "BLOCKING" for row in unresolved_rows)
    target_remaining = sum(row["severity"] == "TARGET_DEPENDENT" for row in unresolved_rows)
    required_fix_remaining = sum(row["severity"] == "REQUIRED_FIX" for row in unresolved_rows)
    blocker_status_consistent = all(
        (row["gate12bl_status"] in {"RESOLVED", "RESOLVED_INHERITED"})
        == (row["gate12bl_resolved"] == "True")
        for row in blocker_rows
    )
    add_check(checks, "BLOCKERS_RESOLVED", "blocker", len(resolved_ids), int(config["expected_resolved_blockers"]), len(resolved_ids) == int(config["expected_resolved_blockers"]) and resolved_ids == EXPECTED_INHERITED_RESOLVED | RESOLVE_IDS)
    add_check(checks, "BLOCKERS_UNRESOLVED", "blocker", len(unresolved_ids), int(config["expected_remaining_blockers"]), len(unresolved_ids) == int(config["expected_remaining_blockers"]) and unresolved_ids == EXPECTED_UNRESOLVED)
    add_check(checks, "BLOCKING_REMAINING", "blocker", blocking_remaining, int(config["expected_blocking_remaining"]), blocking_remaining == int(config["expected_blocking_remaining"]))
    add_check(checks, "TARGET_REMAINING", "blocker", target_remaining, int(config["expected_target_dependent_remaining"]), target_remaining == int(config["expected_target_dependent_remaining"]))
    add_check(checks, "REQUIRED_FIX_REMAINING", "blocker", required_fix_remaining, 0, required_fix_remaining == 0)
    add_check(checks, "BI07_REMAINS", "blocker", next(row["gate12bl_status"] for row in blocker_rows if row["blocker_id"] == "BI-07"), "UNRESOLVED", next(row["gate12bl_status"] for row in blocker_rows if row["blocker_id"] == "BI-07") == "UNRESOLVED")
    add_check(checks, "BLOCKER_STATUS_CONSISTENCY", "blocker", blocker_status_consistent, True, blocker_status_consistent)

    hard_failures = [row for row in checks if row["severity"] == "HARD" and row["status"] != "PASS"]
    if hard_failures:
        write_tsv(
            staging / "admin/GATE12BL_CHECKS.tsv",
            checks,
            ["check_id", "domain", "observed", "expected", "status", "severity", "note"],
        )
        raise RuntimeError("Gate12BL hard checks failed: " + ", ".join(row["check_id"] for row in hard_failures))

    blocker_fields = list(blocker_rows[0])
    write_tsv(staging / "admin/GATE12BL_BLOCKER_DISPOSITION.tsv", blocker_rows, blocker_fields)
    change_rows = [
        {"change_id": "BL-01", "artifact": manuscript_path.relative_to(staging).as_posix(), "location": "Data availability", "action": "REPLACED", "evidence": "Removed stale internal label and distinguished the included release-build audit configuration from absent scientific-analysis materials.", "scientific_result_changed": False},
        {"change_id": "BL-02", "artifact": manuscript_path.relative_to(staging).as_posix(), "location": "Results/ecotype paragraph", "action": "ADDED_CROSS_REFERENCE", "evidence": S4_RESULTS_SENTENCE, "scientific_result_changed": False},
        {"change_id": "BL-03", "artifact": legends_path.relative_to(staging).as_posix(), "location": "Heading", "action": "RENAMED", "evidence": "Replaced inherited internal Gate12BH heading with gate-neutral Figure legends heading.", "scientific_result_changed": False},
        {"change_id": "BL-04", "artifact": "tables/Supplementary_Tables_S4_and_S6.xlsx", "location": "new artifact", "action": "CREATED", "evidence": "Formal workbook is cell-identical to frozen S4/S6 TSV data regions.", "scientific_result_changed": False},
        {"change_id": "BL-05", "artifact": "tables/Supplementary_Tables_S4_and_S6.pdf", "location": "new artifact", "action": "CREATED", "evidence": "Two-page journal-neutral visual-review copy.", "scientific_result_changed": False},
        {"change_id": "BL-06", "artifact": "tables/Supplementary_Table_Legends.md", "location": "new artifact", "action": "CREATED", "evidence": "Standalone titles, scope descriptions, footnotes and numbering warning.", "scientific_result_changed": False},
    ]
    write_tsv(
        staging / "admin/GATE12BL_CHANGE_AUDIT.tsv",
        change_rows,
        ["change_id", "artifact", "location", "action", "evidence", "scientific_result_changed"],
    )
    write_tsv(
        staging / "admin/GATE12BL_CHECKS.tsv",
        checks,
        ["check_id", "domain", "observed", "expected", "status", "severity", "note"],
    )

    input_hash_rows = [
        {
            "role": role,
            "relative_path": path.relative_to(root).as_posix(),
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
            "expected_sha256": expected_hashes[role],
            "status": "PASS",
        }
        for role, path in paths.items()
    ]
    input_hash_rows.append(
        {
            "role": "run_config",
            "relative_path": config_path.relative_to(root).as_posix(),
            "bytes": config_path.stat().st_size,
            "sha256": sha256_file(config_path),
            "expected_sha256": "recorded_at_execution",
            "status": "PASS",
        }
    )
    write_tsv(
        staging / "provenance/GATE12BL_INPUT_HASHES.tsv",
        input_hash_rows,
        ["role", "relative_path", "bytes", "sha256", "expected_sha256", "status"],
    )
    environment_path = staging / "provenance/GATE12BL_ENVIRONMENT.txt"
    environment_path.write_text(
        "\n".join(
            [
                f"python={sys.version.replace(os.linesep, ' ')}",
                f"platform={platform.platform()}",
                f"executable={sys.executable}",
                f"cwd={root}",
                f"soffice={shutil.which('soffice')}",
                f"timezone={started.tzname()}",
                f"openpyxl={importlib.metadata.version('openpyxl')}",
                f"reportlab={importlib.metadata.version('reportlab')}",
                f"pypdf={importlib.metadata.version('pypdf')}",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    report_path = staging / "GATE12BL_REPORT.md"
    report_path.write_text(
        "\n".join(
            [
                "# Gate12BL reproducibility-aligned release report",
                "",
                "## Verdict",
                "",
                "- Package truthfulness: **PASS**",
                "- Supplementary-table integrity: **PASS**",
                "- Scientific-core invariance: **PASS**",
                "- Submission readiness: **NOT_READY**",
                "",
                "## Completed scope",
                "",
                "- BI-10 resolved: no internal Gate label remains in Data availability.",
                "- BI-11 resolved: the active text explicitly states that analysis/rendering scripts are not included.",
                "- BI-12 resolved: the active text distinguishes the included release-build audit configuration from absent scientific-analysis configurations and intermediate objects.",
                "- BI-13 resolved: formal XLSX/PDF tables and standalone legends are present and match the frozen TSV data.",
                f"- Inherited payload: {len(source_figure_files)} figures and {len(source_data_files)} source-data files; all hashes preserved.",
                "",
                "## Remaining boundary",
                "",
                f"- Unresolved blockers: {len(unresolved_rows)} ({blocking_remaining} blocking; {target_remaining} target-dependent).",
                "- BI-07 remains blocking because no public repository DOI or stable URL was created.",
                "- Author list, affiliations, corresponding author, funding, conflicts and CRediT roles remain author-controlled.",
                "- Editable manuscript format, AI disclosure adaptation and journal compliance remain target-dependent.",
                "- Supplementary-table numbering gaps are preserved for the target-journal formatting stage.",
                "",
                "## Active files",
                "",
                f"- Manuscript: `{manuscript_path.relative_to(staging).as_posix()}`",
                f"- Figure legends: `{legends_path.relative_to(staging).as_posix()}`",
                "- Workbook: `tables/Supplementary_Tables_S4_and_S6.xlsx`",
                "- PDF tables: `tables/Supplementary_Tables_S4_and_S6.pdf`",
                "- Table legends: `tables/Supplementary_Table_Legends.md`",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    ended = datetime.now().astimezone()
    run_log_path = staging / "GATE12BL_RUN_LOG.txt"
    run_log_path.write_text(
        "\n".join(
            [
                "gate=Gate12BL",
                f"command={run_command}",
                f"working_directory={root}",
                f"started_at={started.isoformat()}",
                f"ended_at={ended.isoformat()}",
                "exit_code=0",
                f"hard_checks={sum(row['severity'] == 'HARD' for row in checks)}",
                f"hard_checks_passed={sum(row['severity'] == 'HARD' and row['status'] == 'PASS' for row in checks)}",
                f"warnings={sum(row['status'] == 'WARN' for row in checks)}",
                f"figures={len(source_figure_files)}",
                f"source_data_files={len(source_data_files)}",
                f"remaining_blockers={len(unresolved_rows)}",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    manifest_path = staging / "provenance/GATE12BL_OUTPUT_MANIFEST.tsv"
    manifest = manifest_rows(
        staging,
        {"provenance/GATE12BL_OUTPUT_MANIFEST.tsv", "GATE12BL_RECEIPT.json"},
    )
    write_tsv(manifest_path, manifest, ["relative_path", "bytes", "sha256"])
    receipt = {
        "schema_version": "1.0",
        "gate_id": "Gate12BL",
        "run_id": "gate12bl_run_v1",
        "status": "COMPLETED",
        "command": run_command,
        "working_directory": str(root),
        "started_at": started.isoformat(),
        "ended_at": ended.isoformat(),
        "exit_code": 0,
        "verdicts": {
            "package_truthfulness": "PASS",
            "supplementary_table_integrity": "PASS",
            "scientific_core_invariance": "PASS",
            "submission_readiness": "NOT_READY",
        },
        "inventory": {
            "figures": len(source_figure_files),
            "source_data_files": len(source_data_files),
            "supplementary_table_data_rows": len(s4_data) + len(s6_data),
            "output_manifest_rows": len(manifest),
        },
        "blockers": {
            "resolved_total": len(resolved_ids),
            "resolved_this_gate": sorted(RESOLVE_IDS),
            "inherited_resolved": sorted(EXPECTED_INHERITED_RESOLVED),
            "remaining_total": len(unresolved_rows),
            "blocking_remaining": blocking_remaining,
            "required_fix_remaining": required_fix_remaining,
            "target_dependent_remaining": target_remaining,
        },
        "checks": {
            "hard_total": sum(row["severity"] == "HARD" for row in checks),
            "hard_passed": sum(row["severity"] == "HARD" and row["status"] == "PASS" for row in checks),
            "warnings": [row["note"] for row in checks if row["status"] == "WARN"],
        },
        "active_manuscript": {
            "path": manuscript_path.relative_to(staging).as_posix(),
            "sha256": sha256_file(manuscript_path),
        },
        "supplementary_tables": {
            "workbook": {"path": workbook_path.relative_to(staging).as_posix(), "sha256": sha256_file(workbook_path)},
            "pdf": {"path": pdf_path.relative_to(staging).as_posix(), "sha256": sha256_file(pdf_path), "pages": len(pdf_reader.pages)},
            "legends": {"path": table_legends_path.relative_to(staging).as_posix(), "sha256": sha256_file(table_legends_path)},
        },
        "output_manifest": {
            "path": manifest_path.relative_to(staging).as_posix(),
            "sha256": sha256_file(manifest_path),
            "rows": len(manifest),
        },
    }
    (staging / "GATE12BL_RECEIPT.json").write_text(
        json.dumps(receipt, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    staging.rename(output)
    print(
        json.dumps(
            {
                "gate": "Gate12BL",
                "package_truthfulness": "PASS",
                "supplementary_tables": "PASS",
                "submission_readiness": "NOT_READY",
                "output": str(output),
            },
            ensure_ascii=False,
        )
    )
    return 0


def main() -> int:
    try:
        return run_gate()
    except Exception:
        if len(sys.argv) == 2:
            try:
                root = Path.cwd().resolve()
                config_path = Path(sys.argv[1]).resolve()
                config = load_config(config_path)
                output = root / config["output_dir"]
                staging = output.parent / f".{output.name}.staging"
                if staging.exists():
                    stamp = datetime.now().astimezone().strftime("%Y%m%dT%H%M%S%z")
                    failed = output.parent / f".{output.name}.failed-{stamp}"
                    staging.rename(failed)
            except Exception:
                pass
        raise


if __name__ == "__main__":
    raise SystemExit(main())
